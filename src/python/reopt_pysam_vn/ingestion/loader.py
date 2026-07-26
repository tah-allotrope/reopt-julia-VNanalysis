"""Generic factory load ingestion: CSV, XLSX, JSON → validated 8760 kW series."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


@dataclass
class FactoryLoadResult:
    loads_kw: list[float]
    cleaning_summary: dict
    source_path: str
    source_format: str
    detected_column: str
    synthesis_method: str = "none"


class LoadLengthError(ValueError):
    def __init__(self, actual_length: int, likely_resolution: str):
        self.actual_length = actual_length
        self.likely_resolution = likely_resolution
        super().__init__(
            f"Expected 8760 hourly values, got {actual_length}. "
            f"Likely resolution: {likely_resolution}"
        )


_LOAD_COLUMN_PATTERNS = [
    "load_kw", "load", "demand", "consumption", "kw",
    "demand_kw", "power", "energy",
]


def _guess_resolution(length: int) -> str:
    resolutions = {
        35040: "15-minute intervals",
        17520: "30-minute intervals",
        8760: "hourly",
        4380: "2-hour intervals",
        2920: "3-hour intervals",
        12: "monthly totals",
        52: "weekly totals",
        365: "daily totals",
    }
    return resolutions.get(length, f"unknown ({length} rows)")


def clean_numeric(raw_value: object) -> Optional[float]:
    if raw_value is None:
        return None
    text = str(raw_value).replace("﻿", "").strip()
    text = text.strip('"').strip()
    text = text.replace(",", "")
    if text in {"", "-", "NA", "N/A", "None", "null"}:
        return None
    return float(text)


def interpolate_missing(
    values: list[Optional[float]],
) -> tuple[list[float], dict]:
    filled = list(values)
    interpolated_indices: list[int] = []

    for index, value in enumerate(filled):
        if value is not None:
            continue

        left_index = index - 1
        while left_index >= 0 and filled[left_index] is None:
            left_index -= 1

        right_index = index + 1
        while right_index < len(filled) and filled[right_index] is None:
            right_index += 1

        left_value = filled[left_index] if left_index >= 0 else None
        right_value = filled[right_index] if right_index < len(filled) else None

        if left_value is None and right_value is None:
            raise ValueError("Load series contains only missing values")
        if left_value is None:
            filled[index] = float(right_value)  # type: ignore[arg-type]
        elif right_value is None:
            filled[index] = float(left_value)
        else:
            filled[index] = (left_value + right_value) / 2.0

        interpolated_indices.append(index)

    return [float(v) for v in filled if v is not None], {
        "missing_count": len(interpolated_indices),
        "interpolated_indices": interpolated_indices,
    }


def sanitize_load_series(
    values: list[Optional[float]],
) -> tuple[list[float], dict]:
    clipped_negative_count = 0
    precleaned: list[Optional[float]] = []

    for value in values:
        if value is not None and value < 0:
            precleaned.append(0.0)
            clipped_negative_count += 1
        else:
            precleaned.append(value)

    filled, issues = interpolate_missing(precleaned)
    cleaned = [max(v, 0.0) for v in filled]

    issues["clipped_negative_count"] = clipped_negative_count
    issues["final_count"] = len(cleaned)
    return cleaned, issues


def _match_column_header(header: str) -> bool:
    normalized = header.lower().strip().replace(" ", "_")
    return any(pattern in normalized for pattern in _LOAD_COLUMN_PATTERNS)


def _detect_load_column_index(headers: list[str]) -> tuple[int, str]:
    for i, header in enumerate(headers):
        if _match_column_header(header):
            return i, header
    for i, header in enumerate(headers):
        return i, header
    raise ValueError("No columns found in input file")


def _read_csv(
    path: Path, column_hint: Optional[str], timestamp_column: Optional[str],
) -> tuple[list[Optional[float]], str]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))

    if not rows:
        raise ValueError(f"Empty CSV file: {path}")

    headers = rows[0]

    if column_hint:
        col_idx = None
        for i, h in enumerate(headers):
            if h.strip().lower() == column_hint.strip().lower():
                col_idx = i
                break
        if col_idx is None:
            raise ValueError(
                f"Column '{column_hint}' not found. Available: {headers}"
            )
        detected = headers[col_idx]
    else:
        col_idx, detected = _detect_load_column_index(headers)

    values = [
        clean_numeric(row[col_idx] if col_idx < len(row) else None)
        for row in rows[1:]
    ]
    return values, detected


def _read_xlsx(
    path: Path,
    column_hint: Optional[str],
    timestamp_column: Optional[str],
    sheet_name: Optional[str] = None,
) -> tuple[list[Optional[float]], str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            matching = [s for s in workbook.sheetnames if s.lower() == sheet_name.lower()]
            if not matching:
                # scan all sheets for one with 8760-row numeric data
                found_sheet = _scan_sheets_for_load(workbook)
                if found_sheet is None:
                    raise ValueError(
                        f"Sheet '{sheet_name}' not found. "
                        f"Available: {workbook.sheetnames}"
                    )
                worksheet = workbook[found_sheet]
            else:
                worksheet = workbook[matching[0]]
        else:
            # multi-sheet: scan all sheets for one with 8760-row numeric columns
            found_sheet = _scan_sheets_for_load(workbook)
            if found_sheet:
                worksheet = workbook[found_sheet]
            else:
                worksheet = workbook[workbook.sheetnames[0]]

        all_rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not all_rows:
        raise ValueError(f"Empty worksheet in: {path}")

    headers = [str(cell) if cell is not None else f"col_{i}" for i, cell in enumerate(all_rows[0])]

    if column_hint:
        col_idx = None
        for i, h in enumerate(headers):
            if h.strip().lower() == column_hint.strip().lower():
                col_idx = i
                break
        if col_idx is None:
            raise ValueError(
                f"Column '{column_hint}' not found. Available: {headers}"
            )
        detected = headers[col_idx]
    else:
        col_idx, detected = _detect_load_column_index(headers)

    values = [clean_numeric(row[col_idx] if col_idx < len(row) else None) for row in all_rows[1:]]
    return values, detected


def _scan_sheets_for_load(workbook) -> Optional[str]:
    """Scan all sheets for one containing 8760-row numeric data."""
    best_sheet = None
    best_score = -1

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = len(rows) - 1 if len(rows) > 1 else 0

        if data_rows < 100:
            continue

        score = 0
        if data_rows == 8760:
            score += 100
        elif data_rows in (35040, 17520):
            score += 50

        headers = [str(cell).lower() if cell else "" for cell in rows[0]] if rows else []
        if any(_match_column_header(h) for h in headers):
            score += 50

        if score > best_score:
            best_score = score
            best_sheet = sheet_name

    return best_sheet


def _read_json(
    path: Path, column_hint: Optional[str], timestamp_column: Optional[str],
) -> tuple[list[Optional[float]], str]:
    with path.open(encoding="utf-8") as handle:
        data = json.load(handle)

    # REopt scenario format: ElectricLoad.loads_kw
    if isinstance(data, dict):
        if "ElectricLoad" in data and "loads_kw" in data["ElectricLoad"]:
            raw = data["ElectricLoad"]["loads_kw"]
            return [clean_numeric(v) for v in raw], "ElectricLoad.loads_kw"

        if "loads_kw" in data:
            raw = data["loads_kw"]
            return [clean_numeric(v) for v in raw], "loads_kw"

        if column_hint and column_hint in data:
            raw = data[column_hint]
            if isinstance(raw, list):
                return [clean_numeric(v) for v in raw], column_hint

        # search for any 8760-length list
        for key, value in data.items():
            if isinstance(value, list) and len(value) == 8760:
                return [clean_numeric(v) for v in value], key

        # nested search one level deep
        for key, value in data.items():
            if isinstance(value, dict):
                for subkey, subvalue in value.items():
                    if isinstance(subvalue, list) and len(subvalue) == 8760:
                        return [clean_numeric(v) for v in subvalue], f"{key}.{subkey}"

    if isinstance(data, list) and len(data) > 0:
        return [clean_numeric(v) for v in data], "root_array"

    raise ValueError(f"Could not find load data in JSON file: {path}")


def ingest_factory_load(
    path: str | Path,
    column_hint: Optional[str] = None,
    timestamp_column: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> FactoryLoadResult:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    suffix = path.suffix.lower()

    if suffix == ".csv":
        raw_values, detected_column = _read_csv(path, column_hint, timestamp_column)
        source_format = "csv"
    elif suffix in (".xlsx", ".xlsm", ".xls"):
        raw_values, detected_column = _read_xlsx(path, column_hint, timestamp_column, sheet_name)
        source_format = "xlsx"
    elif suffix == ".json":
        raw_values, detected_column = _read_json(path, column_hint, timestamp_column)
        source_format = "json"
    else:
        raise ValueError(f"Unsupported file format: {suffix}")

    original_row_count = len(raw_values)
    cleaned, cleaning_summary = sanitize_load_series(raw_values)

    cleaning_summary["original_row_count"] = original_row_count

    synthesis_method = "none"
    if len(cleaned) != 8760:
        from .synthesize import detect_resolution, route_synthesis

        resolution = detect_resolution(len(cleaned))
        if resolution in ("15min", "30min", "monthly"):
            cleaned, synthesis_method = route_synthesis(cleaned, len(cleaned))
            cleaning_summary["synthesis_method"] = synthesis_method
            cleaning_summary["synthesis_source_rows"] = original_row_count
        else:
            raise LoadLengthError(len(cleaned), _guess_resolution(len(cleaned)))

    return FactoryLoadResult(
        loads_kw=cleaned,
        cleaning_summary=cleaning_summary,
        source_path=str(path),
        source_format=source_format,
        detected_column=detected_column,
        synthesis_method=synthesis_method,
    )
